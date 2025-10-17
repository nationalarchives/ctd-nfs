# Create a QA spreadsheet
# 
# Input Columns: filename_1, filename_2, document_type, county, parish, primary_farm_number, additional_farms, farm_name, addressee_title, addressee_individual_name, addressee_group_names, address, owner_title, owner_individual_name, owner_group_names, owner_address, farmer_title, farmer_individual_name, farmer_group_names, farmer_address, acreage, OS_map_sheet, field_info_date, primary_record_date
#    
# Output Columns: Expected Ref, Ref Warnings, Filenames, File Warnings, Type, Type Warnings, Farm Number, Farm Warnings, Farm Name, Farm Name Warnings, Landowner, Landowner Warnings, Farmer, Farmer Warnings, Acreage, OS Sheet Num, OS Warnings, Field Date, Field Date Warnings, Primary Date, Primary Date Warnings
#
# Need to match rows on filenames (before underscore) + farm number

# Functions:
#   load_spreadsheet_data(processing_folder)
#   output_excel(output_file, values) 
#   filename_pattern_check(filename, row_num)
#   reference_pattern_check(ref, row_num)
#   filename_checks(filename1, filename2, type, row_num)
#   doc_type_check(type, row_num)
#   extract_farms(full_csv)
#   generate_references(box_string, primary_farm_string, additional_farm_string, farm_type, row_num, existing_refs)
#   generate_ref(base_ref, existing_refs)
#   generate_farm_number_for_record(county, parish, farm_nums, ref)
#   generate_farm_numbers(county, parish, farm_nums)
#   get_combined_farm_names_by_ref(farm_names)
#   get_combined_owner_details_by_ref(owner_details)
#   get_combined_farmer_details_by_ref(farmer_details, addressee_details)
#   get_combined_details_by_ref(details, components, expected_count = -1)
#   generate_name(title_value, name_value, group_value, addy_value)
#   array_zip(details_array, key = "")
#   dic_merge(dic1, dic2)
#   date_processing(field_date, primary_dates, row_num)
#   date_check(potential_date, row_num)
#   get_similarity_range(values, get_min = True, get_max = True)


#####################
#   To Do:
#
#   Improve name merging
#   Create full range of test spreadsheets


import csv, re, datetime
import data_normalisation as dn
from pathlib import Path
import calendar

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
#from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

def load_spreadsheet_data(processing_folder):
    ''' Processes any csv files in the designated processing folder

        Keyword Arguments:
            processing_folder - string with path to folder
            
        Returns:
            No returns but throws OSError is file can't be read
    '''
    
    try: 
        for file in Path(processing_folder).glob("*.csv"):     
  
            base_filename = Path(file).stem
            print("Processing file:" + str(base_filename))   
                        
            with open(file, newline='') as f:
                values = csv.DictReader(f) 
                farm_values = extract_farms(values)  
                #print(farm_values)                
            
            #print(file)    
            output_excel(Path(processing_folder, "output", base_filename + ".xlsx"), farm_values)
                
    except OSError as e:
        print("Error in data loading: " + e)
        

def output_excel(output_file, values):
    ''' Saves a spreadsheet containing the processed values for QA checking
    
        keyword Arguments:
            output_file - path to output file
            values - nested dictionary of values to be output. The top keys are the unique reference for each farm and nested below that are the values for that farm with the column name as the keys
            
        Outputs:
            No return but throws OSError if the specified file can't be saved. 
    '''
    
    wb = Workbook()
    sheet = wb.active
    
    headings = ["Reference", "Reference Warnings", "Filenames", "Filename Warnings", "Type", "Type Warnings", "Farm Number", "Farm Number Warnings", "Farm Name", "Farm Name Warnings", "Landowner", "Landowner Warnings", "Farmer", "Farmer Warnings", "Acreage", "Acreage Warnings", "OS Sheet Number", "Field Date", "Field Date Warnings", "Primary Date", "Primary Date Warnings"]
    default_widths = [20, 20, 30, 20, 15, 20, 15, 20, 30, 20, 30, 20, 30, 20,15, 20, 15, 15, 20, 15, 20]
    
    for i in range(0, len(headings)):
        column = headings[i]
        col_num = i+1
        
        sheet.cell(1, col_num, column).font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(col_num)].width = default_widths[i]  
        
        row = 2
        for ref in values:
            sheet.cell(row, 1, ref)
            
            farm_values = values[ref]
            #print(ref + ": " + ", ".join(farm_values))

            if column != "Reference" and column in farm_values.keys():
                sheet.cell(row, headings.index(column)+1, ",\n".join(farm_values[column])).alignment = Alignment(wrap_text=True, vertical='top')
            
            sheet.row_dimensions[row].height = None
            row+=1
    
    try:           
        wb.save(output_file)
        print("Saving " + str(output_file))       
    except OSError as e:
        print("Error in saving spreadsheet: " + e)
        

def filename_pattern_check(filename, row_num):
    ''' Checks whether the filename matches the expected format and extracts the components needed for further processing. If the filename does not match the expected pattern then ValueError exception is thrown.
    
    Keyword Arguments:
        filename - string with the filename to be checked
        row_num - string with number of row in spreadsheet
        
    Returns:
        Tuple containing the central part of the filename and the final count at the end of the filename, a warning or raises an ValueError  
    '''
    
    if filename != "":
        if m := re.match(r"^MAF32-(\d*-\d*)( Pt\d*)?_(\d*).tif$", filename):
            ref_component = m.group(1)
            iteration_num = m.group(3)
            return (ref_component, iteration_num, "")
        if m := re.match(r"^MAF32-(\d*-\d*)( Pt\d*)?.tif$", filename):
            ref_component = m.group(1)
            return (ref_component, 0, "Row " + row_num + ": " + filename + " matches expected cover pattern. Error is this is not a cover.")
        elif m := re.match(r"^MAF32-(\d*-\d*).*(\d*)?.tif$", filename):
            ref_component = m.group(1)
            iteration_num = m.group(2)
            return (ref_component, iteration_num, "Row " + row_num + ": " + filename + " does not match expected pattern. Provisional values have been extracted to use in the reference but their accuracy cannot be guaranteed.")                       
        else:   
            raise ValueError("Row " + row_num + ": " + filename + " does not match expected pattern. Further checks on filenames could not be carried out and an accurate reference could not be generated.")
    else:
        raise ValueError("Row " + row_num + ": Error - Blank filename found. Further checks on the filename could not be carried out and an accurate reference could not be generated.")        


def reference_pattern_check(ref, row_num):
    ''' Checks whether the reference identifier matches the expected pattern
 
    Keyword Arguments:
        filename - string with the filename to be checked
        row_num - string with number of row in spreadsheet
        
    Returns:   
        The reference value or raises a ValueError   
    '''
    
    if ref != "":
        if re.match(r"^MAF 32/\d*/\d*/(Cover|Other)?\d*$", ref):
            #print(ref + " matches pattern.")
            return (ref)
        else:
            #print(ref + " not match pattern.")
            raise ValueError("Row " + row_num + ": " + ref + " does not match expected pattern. Further work with this reference may contain inaccuracies.")
    else:
        raise ValueError("Row " + row_num + ": Error - Blank reference found. Further work with this reference could not be carried out and farm information could not be generated.")        
        
         
# Group 1: Filenames 
# Input columns: A (filename_1) and B (filename_2). 
# Expecting two to five rows of data (depending on number of forms for that farm) - TO DO
# Check 1: the format should be quite consistent: MAF32-\d*-\d*__\d*.tif
# Check 2: the filenames should all match up to the underscore
# Check 3: filename_1, filename_2 should consist of consecutive numbers in each row
# Return comma separated list of file names [Output column: A (Filenames)]
# Warnings if checks don't pass [Output column: B (File Warnings)]
# Warnings if unexpected number of rows matched
def filename_checks(filename1, filename2, type, row_num):
    ''' Carries out the required checks on the filenames for each row
    
        Keyword Arguments:
            filename1 - string with name of first file
            filename2 - string with name of second file
            type - the name of the form
            row_num - string with number of row in spreadsheet 
            
        Returns:
            Tuple containing the central part of the filename for use in the reference and a set of warnings
    '''
    warnings = set()
    
    # check one - match MAF32-\d*-\d*_\d*.tif
    
    if not(filename1 == "" and filename2 == ""):   
        try:
            ref_part1, iteration_num1, warning = filename_pattern_check(filename1, row_num)
            #print(row_num + ": " + ref_part1 + ", warning: " + warning)
            if warning != "":
                warnings.add(warning)
        except ValueError as e:
            warnings.add(str(e))
            ref_part1 = ""
            iteration_num1 = -1
            #print(iteration_num1 + ": " + iteration_num2)
            #print(warnings) 
        
        try:
            if filename2 != "":
                ref_part2, iteration_num2, warning = filename_pattern_check(filename2, row_num)
                if warning != "":
                    warnings.add(warning)
            else:
                ref_part2 = ""
                iteration_num2 = ""

            if type != "Cover" and (type == "Other" and ref_part2 != ""):
                # check two - centre sections match
                if ref_part1 != ref_part2 and ref_part1 != "":
                    warnings.add("Row " + row_num + ": Mismatch in the file names: " + filename1 + ", " + filename2)
                
                # check three - sequence
                if (iteration_num1 != "" and iteration_num2 != "") and (int(iteration_num1) != int(iteration_num2) + 1) and (int(iteration_num1) != int(iteration_num2) - 1) and int(iteration_num1) > 0:
                    warnings.add("Row " + row_num + ": File names (" + iteration_num1 + " and " + iteration_num2 + ") are not consecutive")
            
            # check four - number of filenames
            if (ref_part1 == "" or ref_part2 == "") and type != "Cover" and type != "Other":
                warnings.add("Row " + row_num + ": Only one file name given (" + iteration_num1 + iteration_num2 + ")")                    
            elif type == "Cover" and ref_part2 != "":
                warnings.add("Row " + row_num + ": Type is cover but two file names given (" + iteration_num1 + iteration_num2 + ")")
 

            #print(iteration_num1 + ": " + iteration_num2)
            #print(warnings)  
                
            if ref_part1 != "":
                return (ref_part1, warnings)
            else:
                return (ref_part2, warnings)            
        
        except ValueError as e:
            warnings.add(str(e))
            return ("0-0", warnings)
    else:
        warnings.add("Row " + row_num + ": Error - File names missing")
        return ("0-0", warnings)
        
    
# Group 2: Type of documents
# Input column: C (document_type)
# Expecting two to five rows of data (depending on number of forms for that farm)
# Check 1: allowed values are: "C 51/SSY form", "B 496/EI form", "C 47/SSY form", "C 49/SSY form", "SF form C 69/SSY", "other"
# Return comma separated list of forms [Output column: C (Type)]
# Warnings if content doesn't match allowed values [Output column: D (Type Warnings)]
# Warnings if unexpected number of rows matched

def doc_type_check(type, row_num):
    ''' Checks if the type of document matches one of the valid types (B496/EI, C 47/SSY, C 49/SSY, C51/SSY, SF, SF C69/SSY, Other, Cover)
    
        Keyword Arguments:
            type - string with document type 
            row_num - string with number of row in spreadsheet        
             
        Returns: 
            Type value as string or raises a ValueError
    '''
    
    allowed_types = ["B496/EI", "C 47/SSY", "C 49/SSY", "C51/SSY", "SF", "SF C69/SSY", "Other", "Cover"]
        
    if type in allowed_types:
        return type
    else:
        raise ValueError("Row " + row_num + ": Unknown document type (" + type + ") found.")
    

def extract_farms(full_csv):
    ''' Extracts the data for each farm from the CSV data, checks it and returns the collated data and warnings for each farm (by reference)
    
        Keyword Arguments:
            full_csv - dictionary with the values from the CSV and the column headings as keys
    
        Returns:
            Farm values and warnings in a nested dictionary using the farm reference as a keys and the output column names as the second level keys
    '''
    
    farms = {}
    row_counts = {}
    raw_farm_info = {}    
    
    for row_num, row in enumerate(full_csv, 2):
        file1 = row['filename_1']
        file2 = row['filename_2']
        form = row['document_type']
        ref_component, filename_warnings = filename_checks(file1.strip(), file2.strip(), form, str(row_num))
                
        primary_farm_number = row['primary_farm_number']
        additional_farm_number = row['additional_farms']

        form = form.strip()
        
        type_warnings = set()
        try:
            doc_type_check(form, str(row_num))
        except ValueError as ve:
            type_warnings.update([str(ve)])

        farm_refs, ref_warnings = generate_references(ref_component.replace("-","/"), primary_farm_number.strip(), additional_farm_number.strip(), form, str(row_num), farms.keys())
        
        #print("Row: " + str(row_num) + ", Farm refs (" + ref_component.replace("-","/") + ", farm_num: " + primary_farm_number.strip() + ") :")
        #print(farm_refs)
        #print()
        #print(ref_warnings)

        for temp_ref in farm_refs.keys():
            #core_ref = temp_ref.split("-")[0]
            ref = temp_ref.split("-")[0]
            #print("Checking: " + core_ref + ": " + form)
            if ref in farms.keys() and form not in farms[ref]["Type"]: 
                #ref = core_ref            
                farms[ref]["Type"] += [form]
                #print("Row " + str(row_num) +  ": Core ref in dict, form not in dict. Adding " + form + " to " + ref)       
            elif ref in farms.keys() and form in farms[ref]["Type"]:
                #ref = temp_ref
                #print("Row " + str(row_num) + ": Core ref (" + ref + ") in dict and form in dict")
                if ref in farms.keys():
                    farms[ref]["Type"] += [form]
                else:
                    farms[ref] = {"Type": [form]}  
                                 
                type_warnings.add("Row " + str(row_num) + ": Warning - Multiple '" + form + "' forms for " + ref)
            else:
                #ref = core_ref 
                farms[ref] = {"Type": [form]}
                if ref not in raw_farm_info.keys():
                    raw_farm_info[ref] = dict()
                #print("Row " + str(row_num) + ": Neither Core ref or form in dict. Adding " + form + " to " + ref)

            # count rows
            if ref in row_counts.keys():
                row_counts[ref]["Total"] += 1
            else:
                row_counts[ref] = {"Total":1, "Row": row_num}
        
            # generate farm number    
            county = row['county']
            parish = row['parish']
                           
            if farm_refs[temp_ref] == "Primary":
                farm_numbers = row['primary_farm_number']     
            else:
                farm_numbers = row['additional_farms']
            
            farm_nums = generate_farm_number_for_record(county, parish, farm_numbers, ref)

            if "Farm number" in farms[ref].keys():
                farms[ref]["Farm Number"].update(farm_nums)
                if len(farm_nums) == 1:
                    row_counts[ref]["Farm Number"] += 1
            else:
                farms[ref]["Farm Number"] = farm_nums
                if len(farm_nums) == 1:
                    row_counts[ref].update({"Farm Number":1})

            
            # Type counts            
            if "Type" in row_counts[ref].keys() and form != "":
                row_counts[ref]["Type"] += 1
            elif form != "":
                row_counts[ref].update({"Type":1})                       
        
            # Filenames
            if "Filenames" in farms[ref].keys():
                if file1 != "" and file2 != "":
                    farms[ref]["Filenames"].update({file1, file2})  
                    row_counts[ref]["Filenames"] += 1
                elif file1 != "":
                    farms[ref]["Filenames"].update({file1})  
                else:
                    farms[ref]["Filenames"].update({file2})  
            else:                
                if file1 != "" and file2 != "":
                    farms[ref]["Filenames"] = {file1, file2} 
                    row_counts[ref].update({"Filenames":1})
                elif file1 != "":
                    farms[ref]["Filenames"] = {file1}
                    row_counts[ref].update({"Filenames":0}) 
                else:
                    farms[ref]["Filenames"] = {file2}
                    row_counts[ref].update({"Filenames":0}) 
                
            if "Filename Warnings" in farms[ref].keys():
                farms[ref]["Filename Warnings"].update(filename_warnings)
            else:
                farms[ref]["Filename Warnings"] = filename_warnings
                
            # H (farm_name)
            farm_name = row['farm_name']
            if farm_name != "" and farm_name != "*":
                if "Farm Name" in raw_farm_info[ref].keys():
                    raw_farm_info[ref]["Farm Name"] += [farm_name]

                else:
                    raw_farm_info[ref].update({"Farm Name": [farm_name]})

            
            # M (owner_title), N (owner_individual_name), O (owner_group_names - semi-colon separated list), P (owner_address - semi-colon separated list)
            owner_title = row['owner_title']
            owner_individual_name = row['owner_individual_name']
            owner_group_names = row['owner_group_names']
            owner_addresses = row['owner_address']
            
            combined = owner_title + owner_individual_name + owner_group_names + owner_addresses
            combined = combined.replace("*", "")
            
            # If N/'individual name' contains asterisk, data from O/'group names) should be taken
            
            if len(combined) > 0:
                if "Landowner" in raw_farm_info[ref].keys():
                    raw_farm_info[ref]["Landowner"]["Title"] += [owner_title]
                    raw_farm_info[ref]["Landowner"]["Individual Name"] += [owner_individual_name]
                    raw_farm_info[ref]["Landowner"]["Group Names"] += [owner_group_names.split(";")]
                    raw_farm_info[ref]["Landowner"]["Addresses"] += [owner_addresses.split(";")]
                    
                else:
                    raw_farm_info[ref].update({"Landowner": {"Title": [owner_title]}})
                    raw_farm_info[ref]["Landowner"].update({"Individual Name": [owner_individual_name]})
                    raw_farm_info[ref]["Landowner"].update({"Group Names": [owner_group_names.split(";")]})
                    raw_farm_info[ref]["Landowner"].update({"Addresses": [owner_addresses.split(";")]}) 
                    
                #print(raw_farm_info[ref]["Landowner"])                                

            # Input columns: I (addressee_title), J (addressee_individual_name), K (addressee_group_names), L (address) 
            # Input columns: Q (farmer_title), R (farmer_individual_name), S (farmer_group_names), T (farmer_address)
            addressee_title = row['addressee_title']
            addressee_individual_name = row['addressee_individual_name']
            addressee_group_names = row['addressee_group_names']
            addresses = row['address']
            farmer_title = row['farmer_title']
            farmer_individual_name = row['farmer_individual_name']
            farmer_group_names = row['farmer_group_names']
            farmer_addresses = row['farmer_address']
            
            #combined = addressee_title + addressee_individual_name + addressee_group_names + addresses + farmer_title + farmer_individual_name + farmer_group_names + farmer_addresses
            #combined = combined.replace("*", "")
            
            addressee_combined = addressee_title + addressee_individual_name + addressee_group_names + addresses
            farmer_combined = farmer_title + farmer_individual_name + farmer_group_names + farmer_addresses
            addressee_combined = addressee_combined.replace("*", "")
            farmer_combined = farmer_combined.replace("*", "")
            
            #print("Ref: " + ref)
            #print("Farmer: " + farmer_combined)
            #print("Addressee: " + addressee_combined)
            
            if len(farmer_combined) > 0:            
                if "Farmer" in raw_farm_info[ref].keys():
                    raw_farm_info[ref]["Farmer"]["Title"] += [farmer_title]
                    raw_farm_info[ref]["Farmer"]["Individual Name"] += [farmer_individual_name]
                    raw_farm_info[ref]["Farmer"]["Group Names"] += [farmer_group_names.split(";")]
                    raw_farm_info[ref]["Farmer"]["Addresses"] += [farmer_addresses.split(";")]                                     
                else:
                    raw_farm_info[ref].update({"Farmer": {"Title": [farmer_title]}})
                    raw_farm_info[ref]["Farmer"].update({"Individual Name": [farmer_individual_name]})
                    raw_farm_info[ref]["Farmer"].update({"Group Names": [farmer_group_names.split(";")]})
                    raw_farm_info[ref]["Farmer"].update({"Addresses": [farmer_addresses.split(";")]})                 

            if len(addressee_combined) > 0:            
                if "Addressee" in raw_farm_info[ref].keys():
                    raw_farm_info[ref]["Addressee"]["Title"] += [addressee_title]
                    raw_farm_info[ref]["Addressee"]["Individual Name"] += [addressee_individual_name]
                    raw_farm_info[ref]["Addressee"]["Group Names"] += [addressee_group_names.split(";")]
                    raw_farm_info[ref]["Addressee"]["Addresses"] += [addresses.split(";")]                                     
                else:
                    raw_farm_info[ref].update({"Addressee": {"Title": [addressee_title]}})
                    raw_farm_info[ref]["Addressee"].update({"Individual Name": [addressee_individual_name]})
                    raw_farm_info[ref]["Addressee"].update({"Group Names": [addressee_group_names.split(";")]})
                    raw_farm_info[ref]["Addressee"].update({"Addresses": [addresses.split(";")]})  

            
            # Group 7: Acreage
            # Input column: U (acreage - semi-colon separated list)
            # Expecting 1 row of data
            # Return value [Output column: M (Acreage)]
            # Warning if multiple values     
            acreage = row['acreage']
            if acreage != "":
                if "Acreage" in farms[ref].keys():
                    farms[ref]["Acreage"].update({acreage})
                else:
                    farms[ref].update({"Acreage": {acreage}})
                    
                
                if "Acreage" in row_counts[ref].keys():
                    row_counts[ref]["Acreage"] += 1
                else:
                    row_counts[ref].update({"Acreage":1}) 

                        
            # Group 8: OS
            # Input column: V (OS_map_sheet)
            # Excepting 1 row of data
            # Return value [Output column: O (OS Sheet)]             
            OS_map = row['OS_map_sheet']
            if OS_map != "":
                if "OS Sheet Number" in farms[ref].keys():
                    farms[ref]["OS Sheet Number"].update([OS_map])
                    row_counts[ref]["OS_sheet_number"] += 1
                else:
                    farms[ref].update({"OS Sheet Number": {OS_map}})
                    row_counts[ref].update({"OS_sheet_number":1}) 

                #print(farms[ref]["OS Sheet Number"])
            
            # Dates
            field_date = row["field_info_date"]
            primary_date = row["primary_record_date"]
            
            field_date_values, primary_date_values = date_processing(field_date, primary_date, str(row_num))
            checked_field_date, field_date_warnings = field_date_values
            checked_primary_date, primary_date_warnings = primary_date_values
            
            #print("Checked Field Date: " + checked_field_date)
            #print("Checked Primary Date: " + checked_primary_date)
            
            if checked_field_date != "":
                if "Field Date" in farms[ref].keys():               
                    farms[ref]["Field Date"].add(checked_field_date)
                    row_counts[ref]["Field Date"] += 1
                else:
                    farms[ref].update({"Field Date": {checked_field_date}})
                    row_counts[ref].update({"Field Date":1})
            
            if checked_primary_date != "":            
                if "Primary Date" in farms[ref].keys():                
                    farms[ref]["Primary Date"].add(checked_primary_date)
                    row_counts[ref]["Primary Date"] += 1
                else:
                    farms[ref].update({"Primary Date": {checked_primary_date}})
                    row_counts[ref].update({"Primary Date":1})            
            

            try:
                reference_pattern_check(ref, str(row_num))
            except ValueError as ve:
                if "Reference Warnings" in farms[ref].keys():
                    farms[ref]["Reference Warnings"].update([str(ve)])
                else:
                    farms[ref]["Reference Warnings"] = {str(ve)}   
 
                                
            # Warnings
            if "Reference Warnings" in farms[ref].keys():
                farms[ref]["Reference Warnings"].update(ref_warnings)
            else:
                farms[ref]["Reference Warnings"] = ref_warnings
                
            #print(farms[ref]["Reference Warnings"])

            if "Type Warnings" in farms[ref].keys():
                farms[ref]["Type Warnings"].update(type_warnings)
            else:
                farms[ref]["Type Warnings"] = type_warnings
            
            farm_num_warning = {"Row " + str(row_num) + ": Error in generated farm number - lettercode/parish number/farm number mismatch"}    
            if "Farm Number Warnings" in farms[ref].keys() and len(farm_nums) > 1:
                farms[ref]["Farm Number Warnings"].update(farm_num_warning)
            elif len(farm_nums) > 1:
                farms[ref]["Farm Number Warnings"] = farm_num_warning
            else:
                farms[ref]["Farm Number Warnings"] = set()

            acreage_warning = {"Row " + str(row_num) + ": multiple acreage given"}    
            if "Acreage Warnings" in farms[ref].keys() and ";" in acreage:
                farms[ref]["Acreage Warnings"].update(acreage_warning)
            elif ";" in acreage:
                farms[ref]["Acreage Warnings"] = acreage_warning
            else:
                farms[ref]["Acreage Warnings"] = set()

            if "Field Date Warnings" in farms[ref].keys():
                farms[ref]["Field Date Warnings"].update(field_date_warnings)
            else:
                farms[ref]["Field Date Warnings"] = field_date_warnings

            if "Primary Date Warnings" in farms[ref].keys():
                farms[ref]["Primary Date Warnings"].update(primary_date_warnings)
            else:
                farms[ref]["Primary Date Warnings"] = primary_date_warnings

    #print(raw_farm_info) 
    # Farm names
    farm_names = {}
    for farm_ref, farm_data in raw_farm_info.items():
        if "Farm Name" in farm_data.keys():
            farm_names[farm_ref] = farm_data["Farm Name"]
               
    combined_farm_names, combined_farm_name_warnings = get_combined_farm_names_by_ref(farm_names)
    #print(combined_farm_names)
    
    for ref, combined_farm_name in combined_farm_names.items():
        farms[ref].update({"Farm Name": [combined_farm_name]})
        farms[ref].update({"Farm Name Warnings": combined_farm_name_warnings[ref]})
    
    # Owner names
    owner_details = {}
    for owner_ref, owner_data in raw_farm_info.items():
        if "Landowner" in owner_data.keys():
            owner_details[owner_ref] = owner_data["Landowner"]    
    
    #print(owner_details)
    combined_owner_info, combined_owner_info_warnings = get_combined_owner_details_by_ref(owner_details) 

    for ref, combined_owner_info in combined_owner_info.items():
        farms[ref].update({"Landowner": [combined_owner_info]})
        farms[ref].update({"Landowner Warnings": combined_owner_info_warnings[ref]})    
 
    # Farmer names
    farmer_details = {}
    for farmer_ref, farmer_data in raw_farm_info.items():
        if "Farmer" in farmer_data.keys():
            farmer_details[farmer_ref] = farmer_data["Farmer"]   
    #print("Farmer details: " + str(farmer_details))
    
    # Addressee names
    addressee_details = {}
    for addressee_ref, addressee_data in raw_farm_info.items():
        if "Addressee" in addressee_data.keys():
            addressee_details[addressee_ref] = addressee_data["Addressee"]              
    #print("Addressee details: " + str(addressee_details))
    
    combined_farmer_info, combined_farmer_info_warnings = get_combined_farmer_details_by_ref(farmer_details, addressee_details)
    #print("Combined Farmer Info: " + str(combined_farmer_info))
    
    for ref, combined_farmer_info in combined_farmer_info.items():
        farms[ref].update({"Farmer": [combined_farmer_info]})
        farms[ref].update({"Farmer Warnings": combined_farmer_info_warnings[ref]})   
           
    #print(farms)   
    #print(row_counts)
                
    return (farms)  

    
# Group 0: Reference
# Input columns: A/B (filenames), F (primary_farm_number), G (additional_farms)
# return "MAF 32 " + box number (string after first hyphen and before underscore, replace hyphens with slashed in columns A and B in source which should match) + "/" + farm number (F or G in source) [Output column: U (Reference)] 
# Warnings: if generated using column G [Output column: V (Reference Warnings)] 
def generate_references(box_string, primary_farm_string, additional_farm_string, farm_type, row_num, existing_refs):
    ''' Creates a reference string for each farm in the required format: "MAF 32 " + box number (with slashes) + "/" + farm number
    
        Keyword Arguments:
            box_string - string with the box component of the reference
            primary_farm_string - string with number for primary farm
            additional_farm_string - semi-colon separated list of additional farm numbers
            farm_type - the name of the form
            row_num - string with number of row in spreadsheet
            existing_ref - list of existing references 
            
        Returns:
            Tuple containing a list of generated references and a list of warnings
    '''
    
    ref_list = {}
    warnings = set()
    
    if primary_farm_string == "*":
        ref = generate_ref("MAF 32/" + box_string + "/" + farm_type, existing_refs)
        ref_list[ref] = farm_type
        
    elif primary_farm_string != "" and additional_farm_string == "":
        ref = generate_ref("MAF 32/" + box_string + "/" + primary_farm_string, existing_refs)           
        ref_list[ref] = "Primary"
       
    elif primary_farm_string == "" and additional_farm_string != "":
        for additional_farm in additional_farm_string.split(";"):
            ref = generate_ref("MAF 32/" + box_string + "/" + additional_farm.strip(), existing_refs)           
            ref_list[ref] = "Additional"
        warnings.add("Row " + row_num + ": Error - Additional farm but no primary farm given")
        
    elif primary_farm_string != "" and additional_farm_string != "":
        ref = generate_ref("MAF 32/" + box_string + "/" + primary_farm_string, existing_refs)           
        ref_list[ref] = "Primary"
        
        for additional_farm in additional_farm_string.split(";"):
            ref = generate_ref("MAF 32/" + box_string + "/" + additional_farm.strip(), existing_refs)           
            ref_list[ref] = "Additional"          
               
        warnings.add("Row " + row_num + ": Warning - Additional farms present")
    elif farm_type != "Other" or farm_type != "Cover":        
        warnings.add("Row " + row_num + ": Note - type is " + farm_type.lower() + " so no farm number specified")
        ref = generate_ref("MAF 32/" + box_string + "/" + farm_type, existing_refs)
        ref_list[ref] = farm_type
        
    #print(row_num + ": box - " + box_string + ", primary farm num - " + primary_farm_string + ", type: " + farm_type + ", ref: " + ref)
   
    if type == "Cover" and primary_farm_string != "*":
        warnings.add("Row " + row_num + ": Error - Type is cover and farm number is specified")
        ref = generate_ref("MAF 32/" + box_string + "/" + farm_type, existing_refs)
        ref_list[ref] = farm_type
               
    return (ref_list, warnings)  


def generate_ref(base_ref, existing_refs):
    ''' Check whether the reference already exists and return it or an alternate reference so that the reference is unique
    
        Keyword Arguments: 
            base_ref - string with the proposed reference 
            existing_refs - list of already used references
        
        Returns:
            Reference string      
    '''
    
    if base_ref in existing_refs:
        counter = 1
        temp_ref = base_ref
        while temp_ref in existing_refs:
            temp_ref = temp_ref.split("-")[0] + "-" + str(counter)
            counter += 1   
         
        return temp_ref
    else:
        return base_ref


# Group 3: Farm number
# Input column: D (county), E (parish), F (primary_farm_number) and G (additional_farms - semi-colon separated list) in source spreadsheet
# Expecting three rows of data
# Check 1: lettercode/parish number/farm number they should all match
# Return farm number  [Output column: E (Farm Number)]
# Return additional farm number(s) on separate row(s) and fill in the rest of values from the form that mentions the additional farm(s)
# Warning: generate a warning if not match [Output column: F (Farm Warnings)]
# Warning: flag any case in which column G is filled [Output column: F (Farm Warnings)]
# Warnings if unexpected number of rows matched

def generate_farm_number_for_record(county, parish, farm_nums, ref):
    ''' Generates a unique farm number from the county code, parish code and farm number for a given reference
    
        Keyword Arguments:
            county - string with the county information
            parish - string with the parish information
            farm_nums - string with semi-colon separated list of farm numbers
            ref - string with unique record string
            
        Returns:
            Set with generated farm number (only one number is expected)
    
    '''
    generated_farm_numbers = generate_farm_numbers(county, parish, farm_nums)
    
    farm_generated_number = set()
    
    for generated_number in generated_farm_numbers:
        if ref[-1] == generated_number[-1]:
            farm_generated_number.add(generated_number)       
        
    return farm_generated_number


def generate_farm_numbers(county, parish, farm_nums):
    ''' Generates a unique farm numbers from the county code, parish code and farm number
    
        Keyword Arguments:
            county - string with the county information
            parish - string with the parish information
            farm_nums - string with semi-colon separated list of farm numbers
            
        Returns:
            Set with generated farm numbers    
    '''
    county_code = county.split(" ")[0]
    parish_num = parish.split(" ")[0]
    
    farm_numbers = set()
    
    for farm_num in farm_nums.split(";"):
        farm_numbers.add(county_code + "/" + parish_num + "/" + farm_num)
    
    return farm_numbers
    

# Group 4: Farm
# Input column: H (farm_name)
# Expecting 2 rows of data
# Check 1: values should be similar, ignoring asterisk and blanks. Return combined value.
# Return combined value [Output column: G (Farm Name)]  
# Warning: if not similar [Output column: H (Farm Name Warnings)]
# Warnings if unexpected number of rows matched   

def get_combined_farm_names_by_ref(farm_names):
    ''' Creates a combined value for the farm name from the list of names given for a particular farm
    
        Keyword Arguments: 
            farm_names - dictionary with list of farm names for each farm using the farm reference as the key
            
        Returns:
            Tuple with dictionary of combined name value for each farm, and a dictionary of with a set of warnings for each farm
    '''
    
    combined_names, warnings = dn.component_compare(farm_names)
    
    for ref in farm_names.keys(): 
        count = len(farm_names[ref])
        if count != 2:
            warnings[ref].add("Expected 2 rows of data, Got " + str(count) + ".")
    
    return (combined_names, warnings)


# Group 5: Landowner
# Input columns: M (owner_title), N (owner_individual_name), O (owner_group_names - semi-colon separated list), P (owner_address - semi-colon separated list)
# Expecting 1 row of data (Column C: "B 496/EI form")
# Return combined values [Output column: I (Landowner)]  
#       M and N are to be merged, if both are given. 
#       If N contains asterisk, data from O should be taken
#       P is the address. 
#       O and P should be matched on position in list
# Check 1: same number of tokens in column O (owner group names) and column P (owner addresses) in semi-colon separated lists
# Warning: if N and O both have data [Output column: J (Landowner Warnings)] 
# Warning: if M and O both have data [Output column: J (Landowner Warnings)] 
# Warning: if mismatch with number of owner group names and owner addresses
# Warnings if unexpected number of rows matched

def get_combined_owner_details_by_ref(owner_details):
    ''' Combines owner values and flags warnings if unexpected values found
    
        Key Arguments:  
            owner_details - dictionary with each reference key containing a dictionary with the following: key: 'Title' - list of string values, 'Individual Name' - list of string values, 'Group Names' - list of lists with string values and 'Addresses' - list of lists with string values
            
        Returns:
            Tuple with a dictionary with the combined values for each field by reference and a dictionary with a set of warnings for each reference
    '''
    return get_combined_details_by_ref(owner_details, ["Title", "Individual Name", "Group Names", "Addresses"], 1)


# Group 6: Farmer
# Input columns: I (addressee_title), J (addressee_individual_name), K (addressee_group_names), L (address) 
# Input columns: Q (farmer_title), R (farmer_individual_name), S (farmer_group_names), T (farmer_address)
# Check 1: column I and column Q are similar (individual name titles)
# Check 2: column J and column R are similar (individual name)
# Check 3: column K and column S are similar  (group names)
# Check 4: column T and column L are similar  (addresses)
# Return combined values (name, address) of either I/Q/J/R or K/S and L/T [Output column: I (Farmer)]  
# Warnings: if I/Q/J/R and K/S both given [Output column: I (Farmer Warnings)]  
# Warnings: if not similar [Output column: I (Farmer Warnings)] 

def get_combined_farmer_details_by_ref(farmer_details, addressee_details):
    ''' Combines farmer values and flags warnings if unexpected values found
    
        Key Arguments:  
            farmer_details - dictionary with each reference key containing a dictionary with the following: key: 'Title' - list of string values, 'Individual Name' - list of string values, 'Group Names' - list of lists with string values and 'Addresses' - list of lists with string values
            addressee_details - dictionary with each reference key containing a dictionary with the following: key: 'Title' - list of string values, 'Individual Name' - list of string values, 'Group Names' - list of lists with string values and 'Addresses' - list of lists with string values
                        
        Returns:
            Tuple with a dictionary with the combined values for each field by reference and a dictionary with a set of warnings for each reference
    ''' 

    warnings = {}
    values = {}
    
    # Create combined values by getting similar values for each field
    # call get_combined_details_by_ref with combined values  
    
    shared = set()
    farmer_only = set()
    addressee_only = set()
    
    if farmer_details.keys() != addressee_details.keys():
        shared = set(farmer_details.keys()).intersection(set(addressee_details.keys()))
        farmer_only = set(farmer_details.keys()) - set(addressee_details.keys())
        addressee_only = set(addressee_details.keys()) - set(farmer_details.keys())
    else:
        shared = set(farmer_details.keys())
    
    #print("Farmer details" + str(farmer_details))
    #print("Addressee details" + str(addressee_details))
    #print("Shared: " + str(shared))
    #print("Farmer: " + str(farmer_only))
    #print("Addressee: " + str(addressee_only))
    
    components = ["Title", "Individual Name", "Group Names", "Addresses"]
    
    for ref in list(shared):
        combined_details = {}
        combined_details[ref] = {}
        warnings[ref] = set()
        farmer_title = farmer_details[ref]["Title"]
        farmer_name = farmer_details[ref]["Individual Name"]
        farmer_groups = farmer_details[ref]["Group Names"]
        farmer_addy = farmer_details[ref]["Addresses"]
        
        addressee_title = addressee_details[ref]["Title"]
        addressee_name = addressee_details[ref]["Individual Name"]
        addressee_groups = addressee_details[ref]["Group Names"]
        addressee_addy = addressee_details[ref]["Addresses"]   
        
        #print("\nRef: " + str(ref))
        #print("Farmer name: " + str(farmer_name))
        #print("Farmer addy: " + str(farmer_addy))
        #print("Addressee addy: " + str(addressee_addy))
        
        #TO DO: Need to combine the values into one key not have each as separate keys:values!!!
        addy_dict = dict()
        combined_addy_dict = dict()
        addy_dict["farmer"] = [str(addy[0]) for addy in farmer_addy]
        addy_dict["addressee"] = [str(addy[0]) for addy in addressee_addy]
            
        #print("Addy dict: " + str(addy_dict))   
        
        combined_addy, combined_addy_warnings = dn.component_compare({'farmer': addy_dict['farmer']})
        #print("Combined addy: " + str(combined_addy))
        #print("Combined addy warnings: " + str(combined_addy_warnings['farmer']))
        combined_addy_dict.update(combined_addy)
        #print("combined_addy_dict: " + str(combined_addy_dict))
        warnings[ref].update(combined_addy_warnings['farmer'])
        #print("Warnings: " + str(warnings))
        
        combined_addy, combined_addy_warnings = dn.component_compare({'addressee': addy_dict['addressee']})
        #print("Combined addy: " + str(combined_addy))
        combined_addy_dict.update(combined_addy)
        #print("combined_addy_dict: " + str(combined_addy_dict))        
        warnings[ref].update(combined_addy_warnings['addressee'])
        #print("Warnings: " + str(warnings))
        
        #print("combined_addy: " + str(combined_addy_dict))
        #print("combined_addressee_addy: " + str(combined_addressee_addy))
        
        #combined_addy = dn.component_compare(addy_dict)    
        #print("combined_addy: " + str(combined_addy))
        
        combined_details[ref].update({"Title": farmer_title + addressee_title})
        combined_details[ref].update({"FTitle": farmer_title})
        combined_details[ref].update({"ATitle": addressee_title})
        
        combined_details[ref].update({"Individual Name": farmer_name + addressee_name})  
        combined_details[ref].update({"FIndividual Name": farmer_name})  
        combined_details[ref].update({"AIndividual Name": addressee_name}) 
        
        combined_details[ref].update({"Group Names": farmer_groups + addressee_groups})
        combined_details[ref].update({"FGroup Names": farmer_groups})
        combined_details[ref].update({"AGroup Names": addressee_groups})
        
        combined_details[ref].update({"Addresses": [[combined_addy_dict["farmer"]], [combined_addy_dict["addressee"]]]}) 
        combined_details[ref].update({"FAddresses": [combined_addy_dict["farmer"]]}) 
        combined_details[ref].update({"AAddresses": [combined_addy_dict["addressee"]]}) 
        
        #print("Addresses (" + ref + "): " + str( combined_details[ref]["Addresses"]))  
        
        #print("Ref: " + str(ref))
        #print("Details: " + str(combined_details[ref]))
        
        split_details = False
        for component in components:
            threshold = 80
            min = get_similarity_range(combined_details[ref][component], get_max=False)
            #print(component + ": " + str(min))
            if min < threshold:
                warnings[ref].add("Warning: Similarity (" + str(min) + ") below threshold (" + str(threshold) + ") for " + component)   
                split_details = True
                #print(ref + ": Adding similarity warning for " + component)
                
        if split_details:
            #print("For " + ref + " combining values with separated farmer/addressee with: " + str(combined_details))
            component_fields = ["FTitle", "ATitle", "FIndividual Name", "AIndividual Name", "FGroup Names", "AGroup Names",  "FAddresses",  "AAddresses"]
            combined_values, combined_warnings = get_combined_details_by_ref(combined_details, component_fields)
            #print("Combined values: " + str(combined_values))
        else:  
            #print("For " + ref + " combining values with combined farmer/addressee with " + str(combined_details))      
            combined_values, combined_warnings = get_combined_details_by_ref(combined_details, components)
        
        warnings = dic_merge(warnings, combined_warnings)
        values = dic_merge(values, combined_values)
        
        #print("Values: " + str(values))

    
    for ref in list(farmer_only):
        combined_details = {}
        combined_details[ref] = {}
        warnings[ref] = set()
        
        farmer_title = farmer_details[ref]["Title"]
        farmer_name = farmer_details[ref]["Individual Name"]
        farmer_groups = farmer_details[ref]["Group Names"]
        farmer_addy = farmer_details[ref]["Addresses"]
        
        combined_details[ref].update({"Title": farmer_title})
        combined_details[ref].update({"Individual Name": farmer_name})  
        combined_details[ref].update({"Group Names": farmer_groups})
        combined_details[ref].update({"Addresses": farmer_addy})  
        warnings[ref].add("No addressee values given")
        
        #print("\nFarmer only")
        #print("Ref: " + str(ref))
        #print("Farmer name: " + str(farmer_name))
        #print("Farmer groups: " + str(farmer_groups))
        #print("Farmer addy: " + str(farmer_addy))
        #print("values: " + str(values))
        
        combined_values, combined_warnings = get_combined_details_by_ref(combined_details, components)
        #print("Combined Values:" + str(combined_values))
        warnings = dic_merge(warnings, combined_warnings)
        values = dic_merge(values, combined_values)
        #print("Values:" + str(values))
    
    for ref in list(addressee_only):
        combined_details = {}
        combined_details[ref] = {}
        warnings[ref] = set()
        addressee_title = addressee_details[ref]["Title"]
        addressee_name = addressee_details[ref]["Individual Name"]
        addressee_groups = addressee_details[ref]["Group Names"]
        addressee_addy = addressee_details[ref]["Addresses"]   
        
        combined_details[ref].update({"Title": addressee_title})
        combined_details[ref].update({"Individual Name": addressee_name})  
        combined_details[ref].update({"Group Names": addressee_groups})
        combined_details[ref].update({"Addresses": addressee_addy}) 
        warnings[ref].add("No farmer values given") 

        #print("\nAddressee only")
        #print("Ref: " + str(ref))
        #print("Addressee name: " + str(addressee_name))
        #print("Addressee groups: " + str(addressee_groups))
        #print("Addressee addy: " + str(addressee_addy))
        #print("values: " + str(values))
        
        combined_values, combined_warnings = get_combined_details_by_ref(combined_details, components)
        #print("Combined Values:" + str(combined_values))
        warnings = dic_merge(warnings, combined_warnings)
        values = dic_merge(values, combined_values)
        #print("Values:" + str(values))
        
    
    #print("Farmer Details:" + str(farmer_details))
    #print("Addressee Details:" + str(addressee_details))
    #print("Combined Details:" + str(combined_details))
    #print("Combined Values:" + str(combined_values))
    #print("Final Values:" + str(values))
    #print("Warnings:" + str(warnings))
    
    #combined_values, combined_warnings = get_combined_details_by_ref(combined_details, components)
    
    
    #print(warnings)
    
    return (values, warnings)
        
    
def get_combined_details_by_ref(details, components, expected_count = -1):
    ''' Combines title, individual name, group name and addressee values as either "title individual name, address" or "group name, address", where there may be multiple group names and addresses, and flags warnings if unexpected values found.
    
        Key Arguments:  
            details - dictionary with each reference key containing a dictionary with one of the following: 
                    reference:  'Title' - list of string values, 
                                'Individual Name' - list of string values, 
                                'Group Names' - list of lists with string values, 
                                'Addresses' - list of lists with string values.
                    reference:  'FTitle' - list of string values,
                                'ATitle' - list of string values,
                                'FIndividual Name' - list of string values,
                                'AIndividual Name' - list of string values,
                                'FGroup Names' - list of string values,
                                'AGroup Names' - list of string values,
                                'FAddresses' - list of string values,
                                'AAddresses' - list of string values.                                           
            components - a list of the expected fields in the details dictionary
            expected_count - optional integer. If positive integer in given then a check is carried out on the number of rows with with data
                        
        Returns:
            Tuple with a dictionary with the combined values for each field by reference, and a dictionary with a set of warnings for each reference
    '''
          
    warnings = {}
    combined_details = {}
    
    titles = {}
    names = {}
    groups = {}
    addresses = {}
    
    for ref, detail in details.items(): 
        warnings[ref] = set()
        #print("\n" + ref)
        #print("get_combined_details_by_ref Initial Values: " + str(detail))
        #print("Components: " + str(components))
        
        '''
        # lists of strings as single value expected
        titles[ref] = "/".join(details["Title"])
        names[ref] = "/".join(details["Individual Name"])
        # lists of lists because multiple values allowed
        for gname in details["Group Names"]:
            if ref in groups.keys():
                groups[ref] += "/".join(gname)
            else:
                groups[ref] = "/".join(gname)
        for addy in details["Addresses"]:
            if ref in addresses.keys():
                addresses[ref] += "/".join(addy)
            else:
                addresses[ref] = "/".join(addy)
        '''

        count = set()
        #for each type of field e.g. Title, Individual Name etc converts values to single string for each
        for component in components:
            #print("Combining: " + component)
            #print("Values: " + str(detail[component]))
            component_length = len(detail[component]) 
            count.add(component_length)
            values_to_merge_dict = {}
            
            if component_length > 1: # Check if there is more that one variation and if so merge
                    
                if "Group Names" in component or "Addresses" in component:
                    #print(detail[component])
                    #print(component)                     
                    values_to_merge_dict = array_zip(detail[component], component)
                    #print("Values to merge: " + str(values_to_merge_dict)) 
                else:
                    values_to_merge_dict[component] = detail[component]
                    #for i, values in enumerate(detail[component]):
                    #        values_to_merge_dict[component+str(i)] = [values]
                
                

                #print("values_to_merge_dict: " + str(values_to_merge_dict)) 
                merged_values, merge_warnings = dn.component_compare(values_to_merge_dict)
                #print("Merged Values: " + str(merged_values))
            
                for warning in merge_warnings.values():
                    if len(warning) > 0:
                        warnings[ref].update(warning)

                if "Title" in component:
                    if ref in titles.keys():
                        titles[ref].update({component: list(merged_values.values())})
                    else:
                        titles[ref] = {component: list(merged_values.values())}
                elif "Individual Name" in component:
                    if ref in names.keys():
                        names[ref].update({component: list(merged_values.values())})
                    else: 
                        names[ref] = {component: list(merged_values.values())}
                elif "Group Names" in component:
                    if ref in groups.keys():
                        groups[ref].update({component: list(merged_values.values())})
                    else:
                        groups[ref] = {component: list(merged_values.values())}
                elif "Addresses" in component:
                    if ref in addresses.keys():
                        addresses[ref].update({component: list(merged_values.values())})
                    else:
                        addresses[ref] = {component: list(merged_values.values())}  
                    #print("Addresses: " + str(addresses[ref]))                 
                else:
                    print("Error in get_combined_details_by_ref: unknown component type: " + component)  
            else: # If there are not multiple variations then get first value
                #print("No variations so no merge needed")
                if "Title" in component:
                    if ref in titles.keys():
                        titles[ref].update({component: detail[component][0]})
                    else:
                        titles[ref] = {component: detail[component][0]}
                elif "Individual Name" in component:
                    if ref in names.keys():
                        names[ref].update({component: detail[component][0]})
                    else: 
                        names[ref] = {component: detail[component][0]}
                elif "Group Names" in component:
                    if ref in groups.keys():
                        groups[ref].update({component: detail[component][0]})
                    else:
                        groups[ref] = {component: detail[component][0]}
                    #print(component + ": " + str(detail[component][0]))
                elif "Addresses" in component:
                    if ref in addresses.keys():
                        if isinstance(detail[component][0], list):
                            addresses[ref].update({component: detail[component][0]})
                        else:
                            addresses[ref].update({component: [detail[component][0]]})
                    else:
                        if isinstance(detail[component][0], list):
                            addresses[ref] = {component: detail[component][0]}
                        else:
                            addresses[ref] = {component: [detail[component][0]]}
                    #print("Addresses: " + str(addresses[ref]))  
                    
    
            
        if len(count) != 1: #Check how many distinct counts are in the set - should only be one as should be the same number of values for each component
            warnings[ref].add("Mismatch in number of expected answers.")    
        
        count_list = list(count)  
        # if a positive value is give for expected_count then check the if there is more than one value in the list or if the first (should be only) value does not match the expected_count 
        if expected_count > 0 and (count_list[0] != expected_count or len(count_list) != 1):
            if len(count_list) != 1:
                warnings[ref].add("Expected " + str(expected_count) + " row of data, Got " + str(count_list.sort()[0]) + "-" + str(count_list.sort()[-1] + "."))             
            else:
                warnings[ref].add("Expected " + str(expected_count) + " row of data, Got " + str(count_list[0]) + ".")
                
                
    for ref in details.keys():       
        if "Addresses" in components:
            # Expected values: ["Title", "Individual Name", "Group Names", "Addresses"]
            name_string, name_warnings = generate_name(titles[ref]["Title"], names[ref]["Individual Name"], groups[ref]["Group Names"], addresses[ref]["Addresses"])
        
            combined_details[ref] = name_string
            warnings[ref].update(name_warnings)
        else:
            # Expected values: ["FTitle", "ATitle" "FIndividual Name", "AIndividual Name", "FGroup Names", "AGroup Names",  "FAddresses",  "AAddresses"]
            
            #print(ref + " get farmer name with title: " + str(titles[ref]["FTitle"]) + ", name: " + str(names[ref]["FIndividual Name"]) + ", group names: " + str(groups[ref]["FGroup Names"]) + ", addresses: " + str(addresses[ref]["FAddresses"]))    
            
            farmer_name_string, farmer_name_warnings = generate_name(titles[ref]["FTitle"], names[ref]["FIndividual Name"], groups[ref]["FGroup Names"], addresses[ref]["FAddresses"])
            #print("Farmer name warnings: " + str(farmer_name_warnings))
            warnings[ref].update(farmer_name_warnings)
            #print(warnings[ref])
            if farmer_name_string != "":
                farmer_name_string = "Farmer: " + farmer_name_string
            #print("Farmer name: " + farmer_name_string)     
            
            addressee_name_string, addressee_name_warnings = generate_name(titles[ref]["ATitle"], names[ref]["AIndividual Name"], groups[ref]["AGroup Names"], addresses[ref]["AAddresses"])
            #print("Addressee name warnings: " + str(addressee_name_warnings))
            warnings[ref].update(addressee_name_warnings)
            #print(warnings[ref])
            
            if addressee_name_string != "":
                addressee_name_string = "Addressee: " + addressee_name_string
                
            #print(ref + " get addressee name with title: " + str(titles[ref]["ATitle"]) + ", name: " + str(names[ref]["AIndividual Name"]) + ", group names: " + str(groups[ref]["AGroup Names"]) + ", addresses: " + str(addresses[ref]["AAddresses"]) + ". Gives: " + addressee_name_string)
            
            if farmer_name_string != "" and addressee_name_string != "":
                combined_details[ref] = farmer_name_string + "/" + addressee_name_string
            elif farmer_name_string != "":
                combined_details[ref] = farmer_name_string
            else:
                combined_details[ref] = addressee_name_string
              
    #print(ref + ": Combined values: " + combined_details[ref]) 
    #print("Warnings: " + str(warnings))          
        
    return(combined_details, warnings)      


def generate_name(title_value, name_value, group_value, addy_value):
    ''' Create a string combining the name components into a single value.

        Key Arguments:  
            title_value - string or list of strings giving the title value(s). Single value expected.
            name_value - string or list of strings giving the individual name value(s). Single value expected.
            group_value - string or list of strings giving the names of group name value(s)
            addy_value - string or list of strings giving the address value(s)
                            
        Returns:
            Tuple with string with the combined value, and a set of warnings   
    
    '''
    
    warnings = set()
    name = ""
    
    #print("Title: " + str(title_value))
    #print("Name: " + str(name_value))
    #print("Group Names: " + str(group_value))
    #print("Addresses: " + str(addy_value))
    
    # Check for expected value types
    if isinstance(title_value, list):
            temp = set(title_value)
            if len(temp) == 1 and (list(temp)[0] == ""):
                title_value = ""
            elif len(temp) == 1:
                title_value = list(temp)[0]                
            else:
                warnings.add("Error: multiple names found when one expected")
                title_value = "/".join(title_value) 
        
    if isinstance(name_value, list):
        temp = set(name_value)
        if len(temp) == 1 and (list(temp)[0] == ""):
            name_value = ""
        elif len(temp) == 1:
            name_value = list(temp)[0]
        else:
            warnings.add("Error: multiple names found when one expected")
            name_value = "/".join(name_value) 

    #if isinstance(group_value, list):
    #    temp = set(group_value)
    #    print(temp)
    #    if len(temp) == 1 and (list(temp)[0] == "" or list(temp)[0] == "*"):
    #        group_value = [group_value[0]]
    
    # Generate Combined String
    if name_value != "*" and name_value != "":
        #name
        if title_value != "*":
            name = title_value + " " + name_value
        else:
            name = name_value

        name = name.strip() 
        
        #print("Name: " + name)
        
        if isinstance(group_value, list) and True in [True for group in group_value if group != "*" and group != ""]:
            warnings.add("Error: values found in both name (" + name_value + ") and groups (" + ";".join(group_value) + ")")    

        # Address            
        address = ""
        if isinstance(addy_value, list) and len(addy_value) > 1:
            warnings.add("Error: multiple addresses found when one expected") 
            address = "/".join(addy_value)  
        elif isinstance(addy_value, list):
            address = addy_value[0]
        else:
            address = addy_value
            
        if address == "*":
            address = "[..?]"
            
              
        combined_string = name + ", " + address  
        #print("Combined name + address: " + combined_string)
        
    elif len(group_value) > 0 and len(addy_value) > 0:  
        group_names = []      
        #print("Groups: " + str(group_value) + ", addys: " + str(addy_value))      
        if isinstance(group_value, list) and isinstance(addy_value, list): 
            if len(group_value) != len(addy_value):
                warnings.add("Error: Length of group names (" + str(len(group_value))+ ") and addresses (" + str(len(addy_value)) + ") different") 
            
            for i, group in enumerate(group_value):
                group = group.strip()
                if group == "*":
                    group = "[..?]"
                if i < len(addy_value) and addy_value[i].strip() != "":
                    addy = addy_value[i]
                    if addy.strip() == "*":
                        addy = "[..?]"
                    group_names.append(group + ", " + addy)
                else:
                    group_names.append(group)  
                    warnings.add("Warning: No address with " + group)   
                    
            if len(addy_value) > len(group_value):
                for i in range(len(group_value) - 1, len(addy_value)):
                    addy = addy_value[i]
                    if addy.strip() == "*":
                        addy = "[..?]"
                    group_names.append("[..?], " + addy)
                    warnings.add("Warning: No group name with " + addy)                
                
        else:
            #print("get_combined_details_by_ref: Expecting lists for group and address values got " + str(group_value) + " and " + str(addy_value))
            warnings.add("Error: Expecting lists for group and address values got " + str(group_value) + " and " + str(addy_value) + ". Not able to process")
            
        combined_string = "; ".join(group_names) 
    else:
        #print("Group value: " + str(group_value) + " and Addy value:" + str(addy_value))
        combined_string = ""  
    
    return combined_string, warnings

def array_zip(details_array, key = ""):
    ''' Takes a dictionary with an array of arrays and zips the arrays together, extending the shorter array if necessary
    
        Key Arguments:
            details_array - dictionary containing and array of arrays in the values
            key - optional string to be added to the key in the returned dictionary
            
        Returns:
            Dictionary with the same keys as the input but with each component in the original lists zipped together
    '''
    
    values_to_merge_dict = {}
    longest = 0  
    
    for part in details_array:
        if len(part) > longest:
            longest = len(part)
    
    for i in range(0, len(details_array)):
        while len(details_array[i]) < longest:
            details_array[i].append("") 
            
    values_to_merge = list(zip(*details_array))
    
    for i, values in enumerate(values_to_merge):       
        values_to_merge_dict[key + str(i)] = list(values)
        
    return values_to_merge_dict


def dic_merge(dic1, dic2):
    ''' Deep merge of two dictionaries
    
        Key Arguments:
            dic1 - nested dictionary 
            dic2 - nested dictionary
            
        Returns:
            Merged dictionary
    '''
    
    merged_dic = {}
    merged_keys = set(dic1.keys()).union(set(dic2.keys()))
    
    for key in merged_keys:
        if key in dic1 and key in dic2:
            if isinstance(dic1[key], dict) and isinstance(dic2[key], dict):
                merged_dic[key] = dic_merge(dic1[key], dic2[key])
            elif isinstance(dic1[key], set) and isinstance(dic2[key], set):  
                merged_dic[key] = dic1[key].union(dic2[key])
            elif isinstance(dic1[key], list) and isinstance(dic2[key], list): 
                merged_dic[key] = dic1[key] + dic2[key]
            else:
                merged_dic[key] = dic1[key]
                merged_dic[key].update(dic2[key])
                
        elif key in dic1:
            merged_dic[key] = dic1[key]
        else:
            merged_dic[key] = dic2[key]
    
    return merged_dic


# Group 9: Field Date
# Input column: W (field_info_date)
# Expecting 1 row of data
# Check if valid date
# Return value [Output column: Q (Field Date)] 
# Warnings: if not valid date [Output column: R (Field Date Warnings)] 
# Warnings: if date is not earlier or equal to Primary Date [Output column: R (Field Date Warnings)] 
# Warnings: if unexpected number of rows matched

# Group 9: Primary Date
# Input column: X (primary_record_date - semi-colon separated)
# Expecting 1 row of data
# Check if valid date
# Return value as comma separated list if more then one date [Output column: S (Primary Date)] 
# Warnings: if not valid date
# Warnings: if date is not later or equal to Field Date [Output column: T (Primary Date Warnings)] 
# Warnings: if unexpected number of rows matched

def date_processing(field_date, primary_dates, row_num):
    ''' Checks if dates are valid and that field date occurred before primary date(s)
    
        Key Arguments: 
            field_date - string with the date of the field information
            primary_dates - semi-colon separated string with the date or dates of the primary records
            row_num - string with the number of the row in the source spreadsheet
            
        Returns:
            Tuple of tuples. The first tuple contains a string with the field date value and a set of warnings related to the field dates, the second tuple contains the same but for the primary dates.
    '''
    
    if not(field_date == "" or field_date == "*"):
        checked_field_date, field_date_warnings = date_check(field_date, row_num)
    else:
        checked_field_date = ""
        field_date_warnings = set()
    
    primary_date_warnings = set()
    primary_date_list = []
    
    if type(checked_field_date) is datetime.datetime:
        checked_field_date_str = checked_field_date.strftime('%d %B %Y')
        #print("Checked field date is datetime. String version is " + checked_field_date_str)
    else:
        checked_field_date_str = checked_field_date
        #print("Checked field date is not datetime. String version is " + checked_field_date_str)
    
    if len(primary_dates.split(";")) > 1:
        primary_date_warnings.add("Row " + row_num + ": Multiple dates listed.")
    
    for primary_date in primary_dates.split(";"):
        if not(primary_date == "" or primary_date == "*"):
            checked_primary_date, pdate_warning = date_check(primary_date, row_num)
            
            if type(checked_primary_date) is datetime.datetime:
                checked_primary_date_str = checked_primary_date.strftime('%d %B %Y')
            else:
                checked_primary_date_str = checked_primary_date
            
            primary_date_warnings.update(pdate_warning)
            primary_date_list += [checked_primary_date_str]            
            
            if type(checked_primary_date) is datetime.datetime and type(checked_field_date) is datetime.datetime and checked_primary_date < checked_field_date:
                date_warning = "Row " + row_num + ": Error - Primary Date (" + checked_primary_date_str + ") earlier than Field date (" + checked_field_date_str + ")"
                field_date_warnings.add(date_warning)
                primary_date_warnings.add(date_warning)
                   
    return ((checked_field_date_str, field_date_warnings), (", ".join(primary_date_list), primary_date_warnings))        
    

def date_check(potential_date, row_num):    
    ''' Checks if the date, given as a string, is a valid date
    
        Key Arguments:
            potential_date - string containing the date value for checking
            row_num - string with the number of the row in the source spreadsheet
            
        Returns:
            Tuple with either the date as a date object or the original string if it isn't a valid date and a set with any warnings
    '''
    
    warnings = set()
    potential_date = potential_date.strip()

    MONTH_NAMES: list = "|".join(list(calendar.month_name)[1:])
    DATE_DDMMMYYYY_RGX = re.compile(fr"""^(?P<day>\d\d?) +(?P<month>{MONTH_NAMES}) +(?P<year>\d\d\d\d)$""")
    DATE_MMMYYYY_RGX = re.compile(fr"""^(?P<month>{MONTH_NAMES}) +(?P<year>\d\d\d\d)$""")
    
    if DATE_DDMMMYYYY_RGX.search(potential_date):
        try:
            datetime.strptime(potential_date, "%d %B %Y")
        except ValueError as ve:
            # ve = "day is out of range for month":
            warning = f"Row {row_num}: Error: Date ({potential_date}) is invalid ({ve}). Further date checks cannot be carried out."
            return (potential_date, warning)   
                    
    if rgxmatch := DATE_DDMMMYYYY_RGX.search(potential_date) or DATE_MMMYYYY_RGX.search(potential_date):
        if rgxmatch['year'] not in ["1941", "1942", "1943"]:
            warning = f"Row {row_num}: Error: Date ({potential_date}) is not recognized as within the expected range."
            return (potential_date, warning)

    year = int(year)
    
    if month.isdigit():
        month_number = int(month)
    else:
        try:
            month_number = datetime.datetime.strptime(month, '%B').month
        except ValueError as ve:
            try:
                month_number = datetime.datetime.strptime(month, '%b').month
            except ValueError as ve:
                month_number = 0    
            
            warnings.add("Row " + row_num + ": Warning - Month (" + month + ") is not in the expected format.") 
                    
    try:
        date = datetime.datetime(year=year,month=month_number,day=int(day))  
         
    except ValueError as ve:
        warnings.add("Row " + row_num + ": Error - Date (" + potential_date + ") is not recognized as a valid date in the expected format (" + str(ve) + "). Further date checks cannot be carried out.")
        date = potential_date
        
    return (date, warnings)
            


def get_similarity_range(values, get_min = True, get_max = True):
    ''' Gets the range of similarities for a given set of values and returns either the highest similarity, the lowest or both
    
        Keyword arguments:
            values - string or list of strings
            get_min - boolean. If True return the lowest level of similarity found between the values
            get_max - boolean. If True return the highest level of similarity found between the values
            
        Outputs:
            Returns either a number representing either the highest or lowest similarity or a tuple with both lowest and highest values
    '''
    
    #print("Min: " + str(get_min))
    #print("Max: " + str(get_max))
    #print("Checking: " + str(values))
    
    flattened_values = []
    for value in values:
        if isinstance(value, str) and (value.strip() and value.strip() != "*"):
            flattened_values.append(value)
        elif isinstance(value, list):
            value = [x for x in value if x.strip() and x.strip() != "*"]
            flattened_values += value
    
    values_set = set(flattened_values)
    
    if len(values_set) == 1:
        min = 100
        max = 100
    else:
        max = 0
        min = 100
        for i, value1 in enumerate(list(values_set)):
            for j, value2 in enumerate(list(values_set)):
                if i != j:
                    ratio = fuzz.ratio(value1, value2)
                    if ratio > max:
                        max = ratio
                    if ratio < min:
                        min = ratio
    
    #print("Min: " + str(min))
    #print("Max: " + str(max))
                        
    if get_min and get_max:
        return (min, max)
    elif get_min:
        return min
    else:
        return max
        
  

processing_folder = "processing"
load_spreadsheet_data(processing_folder)